from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service # 导入 Service 类
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
import time
import yagmail
import re

def send_mail(sender, to, subject, contents):
    smtp = yagmail.SMTP(user=sender, host='smtp.qq.com')
    smtp.send(to, subject=subject, contents=contents)

def clickXPATH(xpath):
    button = driver.find_element(By.XPATH,xpath)
    button.click()

def inputTextXPATH(xpath,text):
    text_label  = driver.find_element(By.XPATH,xpath)
    text_label.send_keys(text)

def clearTextXPATH(xpath):
    text_label  = driver.find_element(By.XPATH,xpath)
    text_label.clear()

def extract_area(text):
    match = re.search(r"\[(.*?)\]", text)
    if match:
        return match.group(1)
    else:
        return None  # 或者返回一个空字符串 ""

send=True



data = openpyxl.Workbook()

sheet = data["Sheet"]

nowTime=time.strftime('%Y-%m-%d')
#filepath="temp "+generate_code(5)+" .xlsx"
filepath3="浙江政府采购网 "+nowTime+"生成 .xlsx"
data.save(filepath3)
filepath=filepath3
sendDatas=[]

# Chrome options for running headless (no GUI)
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

# ChromeDriver 的路径 (使用 apt 安装时，通常在这里)
chrome_driver_path = "/usr/lib/chromium-browser/chromedriver"

# 创建 Service 对象，指定 ChromeDriver 的路径
service = Service(executable_path=chrome_driver_path)

# 初始化 Chrome WebDriver，并传入 Service 对象
driver = webdriver.Chrome(service=service, options=chrome_options)

try:
    driver.get('https://ggzy.zj.gov.cn/jyxxgk/list.html')
    time.sleep(1)


    clickXPATH("/html/body/div[2]/div/div[2]/div[3]/div/div/ul/li[2]/a")
    time.sleep(1)
    inputTextXPATH("/html/body/div[2]/div/div[2]/div[5]/div/div/input","险")
    time.sleep(1.5)
    clickXPATH("/html/body/div[2]/div/div[2]/div[5]/div/button")
    time.sleep(1)
    
    page=driver.find_elements(By.XPATH, '/html/body/div[2]/div/div[3]')
    
    sheet=data.active
        
    sheet.cell(1,1).value="项目地点"
    sheet.cell(1,2).value="项目名称"
    sheet.cell(1,3).value="发布日期"
    sheet.cell(1,4).value="项目链接"
    maxLenTitle=0
    maxLenUrl=0
    lines=2
    #初始化变量，无需更改

    print(f"geting data")
    page=driver.find_elements(By.XPATH, '/html/body/div[2]/div/div[3]/div[2]/div[2]/ul')
    page=page[0].find_elements(By.XPATH, 'li')
    for li in page:
        div=li.find_element(By.XPATH, 'div')
        a=div.find_element(By.XPATH, 'a')
        span=li.find_element(By.XPATH, 'span')

        title=div.text
        url=a.get_attribute("href")

        if extract_area(title):
            position=title
        else:
            position="位置提取失败"

        
        print(f"finded url:{url}")
        day=span.text

        if maxLenTitle<len(title):
            maxLenTitle=len(title)
        if maxLenUrl<len(url):
            maxLenUrl=len(url)

        sheet.cell(lines,1).value=position
        sheet.cell(lines,2).value=title
        sheet.cell(lines,3).value=day
        sheet.cell(lines,4).hyperlink=url
        if nowTime==day:
            sendDatas.append([position,title,day,url])
            
        lines+=1
            
    sheet.column_dimensions["A"].width=25
    sheet.column_dimensions["B"].width=maxLenTitle*2
    sheet.column_dimensions["C"].width=15
    sheet.column_dimensions["D"].width=maxLenUrl

    data.save(filepath)

except Exception as e:
    print(f"An error occurred: {e}")

finally:
    driver.quit()
    print("Browser closed.")

if send:
    password=os.environ.get("EMAILPASSWORD")
    
    mails=[]
    with open("mail.txt","r+") as file:
       mail=file.read().split("\n")
    for m in mail:
        if m!="":
            mails.append(m)
    if len(sendDatas)==0:
        print(time.strftime('%Y-%m-%d %H:%M:%S'),"without information")
    elif password==None or password=="":
        print(time.strftime('%Y-%m-%d %H:%M:%S'),"please enter the password")
    else:
        yagmail.register('930914114@qq.com', password)
        print(time.strftime('%Y-%m-%d %H:%M:%S'),"start create email files")
        sendlist=[nowTime+"生成<br/>本邮件由程序自动生成<br/>"]
        for sendData in sendDatas:
            sendlist.append(sendData[0])
            sendlist.append(sendData[1])
            sendlist.append(sendData[2])
            sendlist.append('<a href="%s">%s</a>'%(sendData[3],sendData[3]))
            sendlist.append("<br/>")
            #print(sendData[0],sendData[1],sendData[3],end="\n\n\n")
        sendlist.append("<br/>")
        sendlist.append(filepath1)
        sendlist.append(filepath2)
                
        send_mail("930914114@qq.com", mails, f"浙江政府采购网 {nowTime}生成", sendlist)
            
        print(time.strftime('%Y-%m-%d %H:%M:%S'),"successful send email:",mails)

print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),"done")

